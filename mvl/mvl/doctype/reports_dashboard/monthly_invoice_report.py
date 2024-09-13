from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from frappe.utils import  formatdate
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
# import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles



@frappe.whitelist()
def download():
	filename = 'Monthly Invoice Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)

	data= ["Preparation of Monthly Invoice for Manpower Supply"]
	ws.append(data)
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 11)
	data1= title1 (args)
	ws.append(data1)
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= 11)
	ws.append(["Date","Inv No","Reference","Legar Name","Sub ledger","Income Leger","Amount","CGST 9%","SGST 9%","IGST","Total","Remarks"])
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=12):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	data = get_data(args)
	for row in data:
		ws.append(row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
	month = datetime.strptime(str(args.end_date),'%Y-%m-%d')
	mon = str(month.strftime('%B') +' '+ str(month.strftime('%Y')))
	data = [mon]
	return data

@frappe.whitelist()
def get_data(args):
	ss = []
	invoice = frappe.db.sql("""SELECT * FROM `tabMonthly Invoice` WHERE from_date = '%s' and to_date = '%s' """%(args.start_date,args.end_date), as_dict=True)
	for i in invoice:
		tot=i.total_amount+i.cgst_payable_amount+i.sgst_payable_amount+i.igst_payable_amount
		ref_status = frappe.get_value("Invoice Name",{'name':i.invoice_name},['travel_allowance','retainer'])
		month = datetime.strptime(str(args.end_date),'%Y-%m-%d')
		mon = str(month.strftime('%B') +' '+ str(month.strftime('%Y')))
		if ref_status[0] == 1:
			ref = 'Reimbursement Of Travelling Expenses'
		elif ref_status[1] == 1:
			ref = "Retainer's Fees " + mon
		elif ref_status[1] == 0 and ref_status[0] == 0:
			ref = "Manpower supply " + mon
		row =[
			formatdate(args.end_date),
			i.name,
			ref,
			i.invoice_name,
			'',
			'',
			i.total_amount,
			i.cgst_payable_amount,
			i.sgst_payable_amount,
			i.igst_payable_amount,
			tot,
			''
		]
		ss.append(row)
	return ss