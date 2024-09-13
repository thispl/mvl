from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
# import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles



@frappe.whitelist()
def download():
	filename = 'Monthly TDS Deduction Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)

	data= title1 (args)
	ws.append(data)
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 6)
	ws.append(["RET NO.","Employee Name","PAN No","Salary","TDS 1%","Net Pay"])
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center

	data1= get_data(args)
	for row in data1:
		ws.append(row)


	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
	month = datetime.strptime(str(args.end_date),'%Y-%m-%d')
	mon = str(month.strftime('%B') +' '+ str(month.strftime('%Y')))
	data = ["TDS Deduction for the Month Of "+mon]
	return data

@frappe.whitelist()
def get_data(args):
	ss = []
	tds = frappe.db.sql("""SELECT * FROM `tabEmployee` WHERE retainer=1 and status='Active' """, as_dict=True)
	for i in tds:
		if frappe.db.exists("Salary Slip",{'employee':i.name,'start_date':args.start_date,'end_date':args.end_date,'docstatus': ['!=', 2],'arrear_slip':args.arrear_slip}):
	 
			tp = frappe.db.get_value("Salary Slip",{'employee':i.name,'start_date':args.start_date,'end_date':args.end_date,'docstatus': ['!=', 2],'arrear_slip':args.arrear_slip},['gross_pay']) or 0
			per = tp * 0.01
			a = tp-per
			row =[
				i.name,
				i.employee_name,
				i.pan_number,
				tp,
				round(per),
				a
			]
			ss.append(row)
	return ss