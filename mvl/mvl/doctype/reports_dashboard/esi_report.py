from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
# import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download():
	filename = 'ESI Report'
	test = build_xlsx_response(filename)
	
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["IP Number","IP Name","Payment Days","Total Monthly Wages","Reason Code for Zero Working Days"," Last Working Day"])
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center

	data1= get_data(args)
	for row in data1:
		ws.append(row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def get_data(args):
	ss = []
	emp = frappe.db.sql("""SELECT * FROM `tabEmployee` WHERE status='Active' """,as_dict=True)
	for i in emp:
		salary_slip_query = """
			SELECT * FROM `tabSalary Slip`
			WHERE start_date = %s AND end_date = %s AND arrear_slip = %s AND employee = %s
		"""
		salary_slip_data = frappe.db.sql(salary_slip_query, (args.start_date, args.end_date, args.arrear_slip, i.name), as_dict=True)

		if salary_slip_data:
			a = salary_slip_data[0]
			pd = a.payment_days or 0
			esi_amount = frappe.get_value('Salary Detail', {'salary_component': "ESI Wages", 'parent': a.name, 'docstatus': ['!=', 2]}, ["amount"])

			if esi_amount:
				esi = round(esi_amount)
				row = [
					i.get('esi_number', ''),
					i.get('employee_name', ''),
					math.ceil(pd),
					esi,
					"",
					""
				]
				ss.append(row)
	return ss
	