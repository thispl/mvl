import frappe
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)


@frappe.whitelist()
def download():
	filename = 'bank_remittance_report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)

	data = get_data(args)
	for row in data:
		ws.append(row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_employees():
	employee = employee = frappe.db.get_all('Employee',{'status':'Active'},['*'])
	return employee

def get_dates(from_date,to_date):
	no_of_days = date_diff(add_days(to_date, 1), from_date)
	dates = [add_days(from_date, i) for i in range(0, no_of_days)]
	return dates

def get_data(args):
	data = []
	employees = get_employees()
	salary_slips = frappe.db.count('Salary Slip', {'start_date': args.start_date, 'docstatus':0}) or 0
	att_ot_reg = frappe.db.count('Attendance and OT Register', {'start_date': args.start_date, 'docstatus': 1,'principal_employer':'Marketing Manpower'}) or 0
	status = int(salary_slips) + int(att_ot_reg)
	net = frappe.db.sql("""select sum(net_pay) as amount from `tabSalary Slip` where start_date = '%s' and end_date ='%s' and docstatus != 2"""%(args.start_date,args.end_date),as_dict=1)[0]
	lun = frappe.db.sql("""select sum(lunch_allowance) as amount from `tabAttendance and OT Register` where start_date ='%s' and end_date ='%s' and docstatus != 2"""%(args.start_date,args.end_date),as_dict=1)[0] or 0
	tra = frappe.db.sql("""select sum(transport_allowance) as amount from `tabAttendance and OT Register` where start_date ='%s' and end_date ='%s' and docstatus != 2"""%(args.start_date,args.end_date),as_dict=1)[0] or 0
	total = net['amount'] + (lun['amount']) + (tra['amount'])
	row1 = ['1','','TB','','',status,format_date(args.end_date),'','UTIB0001374','912020062524719','','','','','arunachalam@mercantileventures.co.in','','','','','',total]
	data.append(row1)
	salary_slips = frappe.db.get_all('Salary Slip',{'start_date':('between',(args.start_date,args.end_date)),'docstatus': 0},['*'])
	for salary in salary_slips:
		if salary.principal_employer == 'Marketing Manpower':
			att_ot_reg = frappe.db.get_all('Attendance and OT Register', {'start_date': ('between',(args.start_date,args.end_date)), 'docstatus': 1,'principal_employer':'Marketing Manpower','employee':salary.employee},['*'])
			for att in att_ot_reg:
				trans = 0
				if type(att.transport_allowance) == float:
					trans = round(att.transport_allowance)
				else:
					trans = 0
				row = ['','','',9,'','',format_date(args.end_date),'','','','','','','',' ','',salary.employee_name,salary.ifsc_code,'',frappe.db.get_value('Employee',salary.employee,'bank_ac_no'),salary.net_pay]
				row1 = ['','','',9,'','',format_date(args.end_date),'','','','','','',''," ",'',salary.employee_name,salary.ifsc_code,'',frappe.db.get_value('Employee',salary.employee,'bank_ac_no'),trans]
				data.append(row)
				data.append(row1)
		elif salary.principal_employer == 'REL':
			att_ot_reg = frappe.db.get_all('Attendance and OT Register', {'start_date': ('between',(args.start_date,args.end_date)), 'docstatus': 1,'principal_employer':'REL','employee':salary.employee},['*'])
			for att in att_ot_reg:
				lunch = 0
				if type(att.lunch_allowance) == float:
					lunch = round(att.lunch_allowance)
					lun = lunch + salary.net_pay
				else:
					lunch = 0
					lun = salary.net_pay
				row1 = ['','','',9,'','',format_date(args.end_date),'','','','','','',''," ",'',salary.employee_name,salary.ifsc_code,'',frappe.db.get_value('Employee',salary.employee,'bank_ac_no'),lun]
				data.append(row1)
		else:
			row1 = ['','','',9,'','',
				format_date(args.end_date),
				'','','','','','','',' ','',
				salary.employee_name,salary.ifsc_code,'',
				frappe.db.get_value('Employee',salary.employee,'bank_ac_no'),
				salary.net_pay]
			data.append(row1)
	return data        


