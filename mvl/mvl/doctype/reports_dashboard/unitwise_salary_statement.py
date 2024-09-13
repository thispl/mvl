
from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale
import xlrd
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles


@frappe.whitelist()
def download():
	filename = 'Principal Employer Wise Salary Statement'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	data1= title1 (args)
	ws.append(data1)
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= len(title(args))+2)
	header_date = title (args)
	ws.append(["Row Labels"] + header_date + ["Total"])
	align_center = Alignment(horizontal='center',vertical='center')
	fill_red = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
	font_red = Font(bold=True,color='FFFFFF')
	thin_border = Border(left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(title(args))+2):
		for cell in header:
			cell.alignment = align_center
			cell.fill = fill_red
			cell.font = font_red
	fill_pur = PatternFill(start_color='eac4d5', end_color='eac4d5', fill_type='solid')
	for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=len(title(args)) + 2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			cell.fill = fill_pur

	border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	

	data = get_data(args)
	for row in data:
		ws.append(row)

	for header in ws.iter_rows(min_row=3, max_row=len(get_data(args))+2, min_col=1, max_col=1):
		for cell in header:
			cell.font = Font(bold=True)
	header_range = ws['A1':ws.cell(row=len(get_data(args))+2, column=len(title(args)) + 2).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin

	fill_blue = PatternFill(start_color='98c1d9', end_color='98c1d9', fill_type='solid')

	for header in ws.iter_rows(min_row=20, max_row=20, min_col=1, max_col=len(title(args)) + 2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = fill_blue

	for header in ws.iter_rows(min_row=len(get_data(args))+1, max_row=len(get_data(args))+1, min_col=1, max_col=len(title(args)) + 2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = fill_blue
	fill_green = PatternFill(start_color='ACDF87', end_color='ACDF87', fill_type='solid')
	for header in ws.iter_rows(min_row=len(get_data(args))+2, max_row=len(get_data(args))+2, min_col=1, max_col=len(title(args)) + 2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = fill_green

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
	month = datetime.strptime(str(args.end_date),'%Y-%m-%d')
	mon = str(month.strftime('%B') +' '+ str(month.strftime('%Y')))
	data = ["Principal Employer Wise Salary Statement Month of "+mon]
	return data

@frappe.whitelist()
def title (args): 
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
	unit_values = [unit.name for unit in pe]
	return unit_values

@frappe.whitelist()
def get_data(args):
	data = []
	row=[]
	sal_comp2 = frappe.db.sql("""SELECT * FROM `tabSalary Component` WHERE type = 'Earning' ORDER BY `order` ASC""", as_dict=True)
	for comp in sal_comp2:
		if comp.order != 0:
			pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
			v = 0
			amt_column = [comp.name]	
			for u in pe:
				v += 1
				amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
						`tabSalary Detail`.salary_component = '%s' and
						`tabSalary Slip`.docstatus != 2 and
						`tabSalary Slip`.arrear_slip = '%s' and
						`tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC') and 
						principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(comp.name,args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
				amt_column.append(amt[0]['amount'] or 0)
			amte = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
					`tabSalary Detail`.salary_component = '%s' and
					`tabSalary Slip`.docstatus != 2 and
					`tabSalary Slip`.arrear_slip = '%s' and
					`tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC') and 
					start_date = '%s' and end_date ='%s'""" %(comp.name,args.arrear_slip,args.start_date,args.end_date), as_dict=True)
			amt_column.append(amte[0]['amount'] or 0)
			row.append(amt_column)			
	amt_column = ["Sub Total(A)"]	
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
	v = 0
	for u in pe:
		v += 1
		amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","Earned Provident Fund","Employee State Insurance","Attendance Bonus","Insurance","Gratuity","Uniform","Leave Encashment")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND principal_employer = '%s'
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name,args.start_date, args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
	amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","Earned Provident Fund","Employee State Insurance","Attendance Bonus","Insurance","Gratuity","Uniform","Leave Encashment")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,args.start_date, args.end_date), as_dict=True)
	amt_column.append(amt[0]['amount'] or 0)
	row.append(amt_column)
	sal_comp2 = frappe.db.sql("""SELECT * FROM `tabSalary Component` WHERE type = 'Earning' ORDER BY `order` ASC""", as_dict=True)
	for comp in sal_comp2:
		if comp.order in [10,11,12,13,14,15,16]:
			pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
			v = 0
			amt_column = [comp.name]	
			for u in pe:
				v += 1
				amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
						`tabSalary Detail`.salary_component = '%s' and
						`tabSalary Slip`.docstatus != 2 and
						`tabSalary Slip`.arrear_slip = '%s' and
						`tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC') and 
						principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(comp.name,args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
				amt_column.append(amt[0]['amount'] or 0)
			amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
						`tabSalary Detail`.salary_component = '%s' and
						`tabSalary Slip`.docstatus != 2 and
						`tabSalary Slip`.arrear_slip = '%s' and
						`tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC') and 
					start_date = '%s' and end_date ='%s'""" %(comp.name,args.arrear_slip,args.start_date,args.end_date), as_dict=True)
			amt_column.append(amt[0]['amount'] or 0)
			row.append(amt_column)
	sal_comp2 = frappe.db.sql("""SELECT * FROM `tabSalary Component` WHERE type = 'Deduction' ORDER BY `order` ASC""", as_dict=True)
	for comp in sal_comp2:
		if comp.order != 0:
			pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
			v = 0
			amt_column = [comp.name]	
			for u in pe:
				v += 1
				amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where 
						`tabSalary Detail`.salary_component = '%s' and
						`tabSalary Slip`.docstatus != 2 and
						`tabSalary Slip`.arrear_slip = '%s' and
						`tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC') and
						principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(comp.name,args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
				amt_column.append(amt[0]['amount'] or 0)
			amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where 
						`tabSalary Detail`.salary_component = '%s' and
						`tabSalary Slip`.docstatus != 2 and
						`tabSalary Slip`.arrear_slip = '%s' and
						`tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC') and
						start_date = '%s' and end_date ='%s'""" %(comp.name,args.arrear_slip,args.start_date,args.end_date), as_dict=True)
			amt_column.append(amt[0]['amount'] or 0)
			row.append(amt_column)
	amt_column = ["Sub Total(B)"]
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
	v = 0
	for u in pe:
		v += 1
		amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ('Earned Provident Fund', 'Employee State Insurance','Attendance Bonus','Insurance','Gratuity','Uniform','Leave Encashment','Provident Fund','Deduction Employee State Insurance','Salary Advance Detection','Professional Tax')
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND principal_employer = '%s'
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
	amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ('Earned Provident Fund', 'Employee State Insurance','Attendance Bonus','Insurance','Gratuity','Uniform','Leave Encashment','Provident Fund','Deduction Employee State Insurance','Salary Advance Detection','Professional Tax')
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,args.start_date, args.end_date), as_dict=True)
	amt_column.append(amt[0]['amount'] or 0)
	row.append(amt_column)
	amt_column = ["Total (A-B)"]
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`ORDER BY `name` ASC""", as_dict=True)
	v = 0
	for u in pe:
		v += 1
		amt1 = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","Earned Provident Fund","Employee State Insurance","Attendance Bonus","Insurance","Gratuity","Uniform","Leave Encashment")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND principal_employer = '%s'
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)[0]
		amt2 = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ('Earned Provident Fund', 'Employee State Insurance','Attendance Bonus','Insurance','Gratuity','Uniform','Leave Encashment','Provident Fund','Deduction Employee State Insurance','Salary Advance Detection','Professional Tax')
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND principal_employer = '%s'
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)[0]
		if not amt1['amount']:
				amt1['amount'] = 0
		if not amt2['amount']:
			amt2['amount'] = 0
		tot = amt1['amount'] - amt2['amount']	
		amt_column.append(tot or 0)
	amt1_tot = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","Earned Provident Fund","Employee State Insurance","Attendance Bonus","Insurance","Gratuity","Uniform","Leave Encashment")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,args.start_date, args.end_date), as_dict=True)[0]
	amt2_tot = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
			`tabSalary Detail`.salary_component IN ('Earned Provident Fund', 'Employee State Insurance','Attendance Bonus','Insurance','Gratuity','Uniform','Leave Encashment','Provident Fund','Deduction Employee State Insurance','Salary Advance Detection','Professional Tax')
			AND `tabSalary Slip`.docstatus != 2
			AND `tabSalary Slip`.arrear_slip = '%s'
			AND `tabSalary Slip`.unit not in ('CHN - Retainers','TTN - Retainers','ADOC')
			AND start_date = '%s'
			AND end_date = '%s'
	""" % (args.arrear_slip,args.start_date, args.end_date), as_dict=True)[0]
	if not amt1_tot['amount']:
			amt1_tot['amount'] = 0
	if not amt2_tot['amount']:
		amt2_tot['amount'] = 0
	tot_amt = amt1_tot['amount'] - amt2_tot['amount']
	amt_column.append(tot_amt or 0)
	row.append(amt_column)
	return row


## Retainer

@frappe.whitelist()
def download_retainer():
	filename = 'Principal Employer Wise Salary Statement Retainer'
	test = build_xlsx_response_retainer(filename)

def make_xlsx_retainer(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	data1= title1_retainer(args)
	ws.append(data1)
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= len(title_retainer(args))+2)
	header_date = title_retainer (args)
	bold_font = Font(bold=True)
	ws.append(["Row Labels"] + header_date + ["Total"])
	align_center = Alignment(horizontal='center',vertical='center')
	fill_red = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
	font_red = Font(bold=True,color='FFFFFF')
	fill_pur = PatternFill(start_color='eac4d5', end_color='eac4d5', fill_type='solid')
	for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=len(header_date) + 2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			cell.fill = fill_pur
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(title_retainer(args))+2):
		for cell in header:
			cell.alignment = align_center
			cell.fill = fill_red
			cell.font = font_red
	
	border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

	data = get_data1(args)
	for row in data:
		ws.append(row)
	fill_blue = PatternFill(start_color='98c1d9', end_color='98c1d9', fill_type='solid')

	for header in ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=len(header_date)+2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = fill_blue

	for header in ws.iter_rows(min_row=12, max_row=12, min_col=1, max_col=len(header_date)+2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = fill_blue
	fill_green = PatternFill(start_color='ACDF87', end_color='ACDF87', fill_type='solid')
	for header in ws.iter_rows(min_row=13, max_row=13, min_col=1, max_col=len(header_date)+2):
		for cell in header:
			cell.font = Font(bold=True)
			cell.fill = fill_green

	for header in ws.iter_rows(min_row=3, max_row=len(get_data1(args))+2, min_col=1, max_col=1):
		for cell in header:
			cell.font = Font(bold=True)
	header_range = ws['A1':ws.cell(row=len(get_data1(args))+2, column=len(header_date) + 2).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


def build_xlsx_response_retainer(filename):
	xlsx_file = make_xlsx_retainer(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1_retainer(args):
	month = datetime.strptime(str(args.end_date),'%Y-%m-%d')
	mon = str(month.strftime('%B') +' '+ str(month.strftime('%Y')))
	data = ["Principal Employer Wise Salary Statement Retainer Month of "+mon]
	return data

@frappe.whitelist()
def title_retainer (args): 
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer` where retainer = 1 ORDER BY `name` ASC""", as_dict=True)
	unit_values = [unit.name for unit in pe]
	return unit_values

@frappe.whitelist()
def get_data1(args):
	data=[]
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`  where retainer = 1 ORDER BY `name` ASC""", as_dict=True)
	v = 0
	amt_column = ["E Gross Pay"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component in ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","PT Refund") and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' 
				and start_date = '%s' and end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)	
	amt_column = ["Attendance Bonus"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component = "Attendance Bonus" and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' and 
				start_date = '%s' and 
				end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)	
	amt_column = ["Uniform"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component = "Uniform" and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)	
	amt_column = ["Insurance"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component = "Insurance" and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)			
	amt_column = ["Sub Total(A)"]	
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`  where retainer = 1 ORDER BY `name` ASC""", as_dict=True)
	v = 0
	amount = 0
	for u in pe:
		v += 1
		amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","Attendance Bonus","Insurance","Uniform")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers")
				AND principal_employer = '%s'
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)
	amt_column = ["Attendance Bonus"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component = "Attendance Bonus" and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)	
	amt_column = ["Uniform"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component = "Uniform" and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)	
	amt_column = ["Insurance"]
	amount = 0	
	for u in pe:
		v += 1
		amt = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as amount from `tabSalary Slip` left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where
				`tabSalary Detail`.salary_component = "Insurance" and
				`tabSalary Slip`.docstatus != 2 and
				`tabSalary Slip`.arrear_slip = '%s' and
				unit in ("CHN - Retainers","TTN - Retainers") and
				principal_employer ='%s' and start_date = '%s' and end_date ='%s'""" %(args.arrear_slip,u.name,args.start_date,args.end_date), as_dict=True)
		amt_column.append(amt[0]['amount'] or 0)
		amount += amt[0]['amount'] or 0
	amt_column.append(amount or 0)
	data.append(amt_column)	
	amt_column = ["TDS"]
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`  where retainer = 1 ORDER BY `name` ASC""", as_dict=True)
	v = 0
	amount = 0
	amount1 = 0
	for u in pe:
		v += 1
		amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND principal_employer = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers") 
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		if amt[0]['amount']:
			amount = (amt[0]['amount']/100) or 0
		else:
			amount = 0
		amt_column.append(int(amount) or 0)
		amount1 += int(amount)
	amt_column.append(int(amount1) or 0)
	data.append(amt_column)	
	amt_column = ["Sub Total(B)"]	
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`  where retainer = 1 ORDER BY `name` ASC""", as_dict=True)
	v = 0
	amount1 = 0
	amount2 = 0
	amount3 = 0
	amount4 = 0
	for u in pe:
		v += 1
		amt1 = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Attendance Bonus","Insurance","Uniform")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND principal_employer = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers") 
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		amount1 = amt1[0]['amount'] or 0
		amount3 += amount1
		amt2 = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND principal_employer = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers") 
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		if amt2[0]['amount']:
			amount2 = int(amt2[0]['amount']/100) or 0
		else:
			amount2 = 0
		amount4 += amount2
		if not amount1:
			amount1 = 0
		if not amount2:
			amount2 = 0
		tot = amount1 + amount2	
		amt_column.append(tot or 0)
	if not amount3:
		amount3 = 0
	if not amount4:
		amount4 = 0
	tot = amount3 + amount4	
	amt_column.append(tot or 0)
	data.append(amt_column)
	amt_column = ["Total(A - B)"]	
	pe = frappe.db.sql("""SELECT * FROM `tabPrincipal Employer`  where retainer = 1 ORDER BY `name` ASC""", as_dict=True)
	v = 0
	amount = 0
	amount1 = 0
	amount2 = 0
	amount3 = 0
	amount4 = 0
	amount5 = 0
	for u in pe:
		v += 1
		amt = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend","Attendance Bonus","Insurance","Uniform")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND principal_employer = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers")
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		amount = amt[0]['amount'] or 0
		amount3 += amount
		amt1 = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Attendance Bonus","Insurance","Uniform")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND principal_employer = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers") 
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		amount1 = amt1[0]['amount'] or 0
		amount4 += amount1
		amt2 = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) AS amount  FROM `tabSalary Slip` LEFT JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE
				`tabSalary Detail`.salary_component IN ("Earned Basic","Earned Dearness Allowance","Earned House Rent Allowance","Earned Washing Allowance","Earned Conveyance Allowance","Earned Medical Allowance","Earned Special Allowance","Earned Other Allowance","Overtime Amount","Stipend")
				AND `tabSalary Slip`.docstatus != 2
				AND `tabSalary Slip`.arrear_slip = '%s'
				AND principal_employer = '%s'
				AND unit in ("CHN - Retainers","TTN - Retainers") 
				AND start_date = '%s'
				AND end_date = '%s'
		""" % (args.arrear_slip,u.name, args.start_date, args.end_date), as_dict=True)
		if amt2[0]['amount']:
			amount2 = int(amt2[0]['amount']/100) or 0
		else:
			amount2=0
		amount5 += amount2
		if not amount1:
			amount1 = 0
		if not amount2:
			amount2 = 0
		if not amount:
			amount = 0
		tot = amount -(amount1 + amount2)	
		amt_column.append(tot or 0)
	if not amount4:
		amount4 = 0
	if not amount5:
		amount5 = 0
	if not amount3:
		amount3 = 0
	tot = amount3 -(amount4 + amount5)	
	amt_column.append(tot or 0)
	data.append(amt_column)
	return data
	