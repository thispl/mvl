# Copyright (c) 2023, veeramayandi.p@groupteampro.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import (
	DATE_FORMAT,
	add_days,
	add_to_date,
	cint,
	comma_and,
	date_diff,
	flt,
	get_link_to_form,
	getdate,
)
from dateutil.relativedelta import relativedelta

class MonthlyInvoice(Document):
	pass

@frappe.whitelist()
def get_end_date(start_date, frequency):
	start_date = getdate(start_date)
	frequency = frequency.lower() if frequency else "monthly"
	kwargs = (
		get_frequency_kwargs(frequency) if frequency != "bimonthly" else get_frequency_kwargs("monthly")
	)
	end_date = add_to_date(start_date, **kwargs) - relativedelta(days=1)
	if frequency == "monthly":
		return dict(end_date=end_date.strftime(DATE_FORMAT))

	else:
		return dict(end_date="")


def get_frequency_kwargs(frequency_name):
	frequency_dict = {
		"monthly": {"months": 1},
		"fortnightly": {"days": 14},
		"weekly": {"days": 7},
		"daily": {"days": 1},
	}
	return frequency_dict.get(frequency_name)

from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)


@frappe.whitelist()
def download():
	filename = 'Monthly Invoice'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	ws.column_dimensions['A'].width = 10
	ws.column_dimensions['B'].width = 30 
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 15
	ws.column_dimensions['E'].width = 7
	ws.column_dimensions['F'].width = 7
	ws.column_dimensions['G'].width = 7
	ws.column_dimensions['H'].width = 7
	ws.column_dimensions['I'].width = 7
	ws.column_dimensions['J'].width = 7
	ws.column_dimensions['K'].width = 7
	ws.column_dimensions['L'].width = 7
	ws.column_dimensions['M'].width = 9
	ws.column_dimensions['N'].width = 7
	ws.column_dimensions['O'].width = 7
	ws.column_dimensions['P'].width = 7
	ws.column_dimensions['Q'].width = 7
	ws.column_dimensions['R'].width = 7
	ws.column_dimensions['S'].width = 7
	ws.column_dimensions['T'].width = 7
	ws.column_dimensions['U'].width = 7
	ws.column_dimensions['V'].width = 7
	ws.column_dimensions['W'].width = 7
	ws.column_dimensions['X'].width = 9
	ws.column_dimensions['Y'].width = 7
	ws.column_dimensions['Z'].width = 7
	ws.column_dimensions['AA'].width = 7
	ws.column_dimensions['AB'].width = 7
	ws.column_dimensions['AC'].width = 7
	ws.column_dimensions['AD'].width = 7
	ws.column_dimensions['AE'].width = 7
	ws.column_dimensions['AF'].width = 7
	ws.column_dimensions['AG'].width = 7
	ws.column_dimensions['AH'].width = 9
	ws.column_dimensions['AI'].width = 7
	ws.column_dimensions['AJ'].width = 7
	ws.column_dimensions['AK'].width = 7
	ws.column_dimensions['AL'].width = 7
	ws.column_dimensions['AM'].width = 7
	ws.column_dimensions['AN'].width = 7
	ws.column_dimensions['AO'].width = 9

	data = get_data(args)
	for row in data:
		ws.append(row)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=41)
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=41)
	ws.merge_cells(start_row=len(get_data(args)), start_column=1, end_row=len(get_data(args)), end_column=4 )
	
	for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=41):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=41):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=41):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=4, max_row=len(get_data(args)), min_col=1, max_col=4):
		for cell in header:
			cell.font = Font(bold=True)	
	for header in ws.iter_rows(min_row=4, max_row=len(get_data(args)), min_col=13, max_col=13):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=4, max_row=len(get_data(args)), min_col=24, max_col=24):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=4, max_row=len(get_data(args)), min_col=34, max_col=34):
		for cell in header:
			cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=4, max_row=len(get_data(args)), min_col=41, max_col=41):
		for cell in header:
			cell.font = Font(bold=True)

	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=41):
		for cell in header:
			cell.border = border_thin	
			cell.alignment = align_center	
	align_center = Alignment(horizontal='center',vertical='center')
	for header in ws.iter_rows(min_row=3, max_row=len(get_data(args)), min_col=1, max_col=41):
		for cell in header:
			cell.border = border_thin	
			cell.alignment = align_center
	for header in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=41):
		for cell in header:
			cell.alignment = Alignment(horizontal='center',vertical='center',text_rotation=90)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_dates(from_date,to_date):
	no_of_days = date_diff(add_days(to_date, 1), from_date)
	dates = [add_days(from_date, i) for i in range(0, no_of_days)]
	return dates

def get_data(args):
	data = []
	date_object = datetime.strptime(args.from_date, '%Y-%m-%d')
	formatted_date = date_object.strftime('%B %Y')
	date_object = datetime.strptime(formatted_date, '%B %Y')
	formatted_date1 = date_object.strftime('%B %Y').upper()
	head = ["Mercantile Ventures Limited"]
	data.append(head)
	invoice = frappe.get_value("Monthly Invoice",{'name':args.name},['invoice_name'])
	hea = ["WAGE DETAILS FOR THE MONTH OF  " + formatted_date1 + " & ANNEXURE TO INVOICE NUMBER " + " " + args.name + " "]
	data.append(hea)
	row = ["Employee ID","Employee Name","Designation","Location","Days Attend","Basic","DA","HRA","WA","CA","SPA","MA","Total","Earned Basic","EDA","EHRA","EWA","ECA","ESPA","EMA","OA","OT Hrs","OT Amt","Gross","EPF 13%","ESI 3.28%","Bonus","Uniform","Insurance","Leave encashment","Other Additions","Sub Total","Service Charge","Total Payable to MVL","DPF 12%","DESI 0.78%","L/W Fund","P/Tax","Salary Advance","Deductions","Net Pay"]
	data.append(row)
	pd = 0
	t_bas = 0
	t_da = 0
	t_hra = 0
	t_wa = 0
	t_ca = 0
	t_spa = 0
	t_ma = 0
	t_tot = 0
	t_ebas = 0
	t_eda = 0
	t_ehra = 0
	t_ewa = 0
	t_eca = 0
	t_espa = 0
	t_ema = 0
	t_oa = 0
	t_ot_hours = 0
	t_ot_amt = 0
	t_egross = 0
	t_epf = 0
	t_esi = 0
	t_ab = 0
	t_uniform = 0
	t_ins = 0
	t_leave = 0
	t_add = 0
	t_mvl = 0
	t_dpf = 0
	t_desi = 0
	t_lf = 0
	t_ptax = 0
	t_intx = 0
	t_od = 0
	t_sad = 0
	t_net = 0
	t_sc = 0
	t_sl = 0
	slip = frappe.get_all("Salary Slip",{'start_date': args.from_date,'end_date':args.to_date,'invoice_name':invoice,'docstatus':('!=',2)} ,['*'])
	for i in slip:
		bas = frappe.db.get_value("Employee",{'name':i.employee} ,['basic']) 
		da = frappe.db.get_value("Employee",{'name':i.employee} ,['dearness_allowance']) 
		hra = frappe.db.get_value("Employee",{'name':i.employee} ,['house_rent_allowance']) 
		wa = frappe.db.get_value("Employee",{'name':i.employee} ,['washing_allowance']) 
		ca = frappe.db.get_value("Employee",{'name':i.employee} ,['conveyance_allowance']) 
		spa = frappe.db.get_value("Employee",{'name':i.employee} ,['special_allowance']) 
		ma = frappe.db.get_value("Employee",{'name':i.employee} ,['medical_allowance']) 
		gross = bas+da+hra+wa+ca+spa+ma
		ebas = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Basic",'parent':i.name},["amount"]) or 0
		eda = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Dearness Allowance",'parent':i.name},["amount"]) or 0
		ehra = frappe.db.get_value('Salary Detail',{'salary_component':"Earned House Rent Allowance",'parent':i.name},["amount"]) or 0
		ewa = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Washing Allowance",'parent':i.name},["amount"]) or 0
		eca = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Conveyance Allowance",'parent':i.name},["amount"]) or 0
		espa = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Medical Allowance",'parent':i.name},["amount"]) or 0
		ema = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Special Allowance",'parent':i.name},["amount"]) or 0
		oa = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Other Allowance",'parent':i.name},["amount"]) or 0
		ot_hours = i.overtime_hours
		ot_amt = frappe.db.get_value('Salary Detail',{'salary_component':"Overtime Amount",'parent':i.name},["amount"]) or 0
		egross = ebas + eda + ehra + ewa + eca + espa + ema + oa + ot_amt
		epf = frappe.db.get_value('Salary Detail',{'salary_component':"Earned Provident Fund",'parent':i.name},["amount"]) or 0
		esi = frappe.db.get_value('Salary Detail',{'abbr':"ESI",'parent':i.name},["amount"]) or 0
		ab = frappe.db.get_value('Salary Detail',{'salary_component':"Attendance Bonus",'parent':i.name},["amount"]) or 0
		uniform = frappe.db.get_value('Salary Detail',{'salary_component':"Uniform",'parent':i.name},["amount"]) or 0
		ins = frappe.db.get_value('Salary Detail',{'salary_component':"Insurance",'parent':i.name},["amount"]) or 0
		leave = frappe.db.get_value('Salary Detail',{'salary_component':"Leave Encashment",'parent':i.name},["amount"]) or 0
		add = frappe.db.get_value('Salary Detail',{'salary_component':"Other Additions",'parent':i.name},["amount"]) or 0
		ser_ch = frappe.db.get_value('Salary Detail',{'salary_component':"Service Charges",'parent':i.name},["amount"]) or 0
		dpf = frappe.db.get_value('Salary Detail',{'salary_component':"Provident Fund",'parent':i.name},["amount"]) or 0 
		desi = frappe.db.get_value('Salary Detail',{'abbr':"D_ESI",'parent':i.name},["amount"]) or 0 
		lf = frappe.db.get_value('Salary Detail',{'salary_component':"L/w Fund",'parent':i.name},["amount"]) or 0 
		ptax = frappe.db.get_value('Salary Detail',{'salary_component':"Professional Tax",'parent':i.name},["amount"]) or 0 
		intx = frappe.db.get_value('Salary Detail',{'salary_component':"Income Tax",'parent':i.name},["amount"]) or 0 
		sad = frappe.db.get_value('Salary Detail',{'salary_component':"Salary Advance Detection",'parent':i.name},["amount"]) or 0 
		row1 = [i.employee,i.employee_name,i.designation,i.department,i.payment_days,int(bas),int(da),int(hra),int(wa),int(ca),int(spa),int(ma),int(gross),int(ebas),int(eda),int(ehra),int(ewa),int(eca),int(espa),int(ema),int(oa),ot_hours,int(ot_amt),int(egross),int(epf),int(esi),int(ab),int(uniform),int(ins),int(leave),int(add),
	int(i.sub_total),
	int(ser_ch),
	int(i.total_payable_to_mvl),
	int(dpf),
	int(desi),
	int(lf),
	int(ptax),
	# int(intx),
	int(sad),
	int(i.total_deduction),
	int(i.rounded_total)
		]
		pd += i.payment_days
		t_bas += bas
		t_da += da
		t_hra += hra
		t_wa += wa
		t_ca += ca
		t_spa += spa
		t_ma += ma
		t_tot += gross
		t_ebas += ebas
		t_eda += eda
		t_ehra += ehra
		t_ewa += ewa
		t_eca += eca
		t_espa += espa
		t_ema += ema
		t_oa += oa
		t_ot_hours += ot_hours
		t_ot_amt += ot_amt
		t_egross += egross
		t_epf += epf
		t_esi += esi
		t_ab += ab
		t_uniform += uniform
		t_ins += ins
		t_leave += leave
		t_add += add
		t_sc =+ ser_ch
		t_sl =+ i.sub_total
		t_mvl += i.total_payable_to_mvl
		t_dpf += dpf
		t_desi += desi
		t_lf += lf
		t_ptax += ptax
		t_intx += intx
		t_sad += sad
		t_od += i.total_deduction
		t_net += i.rounded_total
		data.append(row1)
	row2 = ["Total","","","",int(pd),
	int(t_bas),
	int(t_da),
	int(t_hra),
	int(t_wa),
	int(t_ca),
	int(t_spa),
	int(t_ma),
	int(t_tot),
	int(t_ebas),
	int(t_eda),
	int(t_ehra),
	int(t_ewa),
	int(t_eca),
	int(t_espa),
	int(t_ema),
	int(t_oa),
	int(t_ot_hours),
	int(t_ot_amt),
	int(t_egross),
	int(t_epf),
	int(t_esi),
	int(t_ab),
	int(t_uniform),
	int(t_ins),
	int(t_leave),
	int(t_add),
	int(t_sl),
	int(t_sc),
	int(t_mvl),
	int(t_dpf),
	int(t_desi),
	int(t_lf),
	int(t_ptax),
	# int(t_intx),
	int(t_sad),
	int(t_od),
	int(t_net)
	]
	data.append(row2)
	return data   


